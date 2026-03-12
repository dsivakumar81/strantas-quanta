from __future__ import annotations

import io

from openpyxl import load_workbook

from quanta_api.domain.enums import DocumentType, ProcessingStatus, SubmissionIntent
from quanta_api.domain.models import Attachment, ReaderInventory, SubmissionEnvelope
from quanta_api.domain.repositories import ObjectStore, SubmissionRepository
from quanta_api.services.pdf_tables import extract_pdf_text


class ReaderService:
    def __init__(self, submission_repository: SubmissionRepository, object_store: ObjectStore) -> None:
        self.submission_repository = submission_repository
        self.object_store = object_store

    def parse_submission(self, submission_id: str) -> ReaderInventory:
        submission = self.submission_repository.get(submission_id)
        if submission is None:
            raise KeyError(submission_id)

        intent = self._classify_intent(submission)
        inventory = [self._classify_attachment(item) for item in submission.attachments]
        warnings = []
        if not inventory:
            warnings.append("Submission arrived without attachments")
        if not any(item.document_type == DocumentType.census for item in inventory):
            warnings.append("No census-like attachment detected during reader pass")

        submission.attachments = inventory
        submission.submission_intent = intent
        submission.document_warnings = warnings
        submission.processing_status = ProcessingStatus.parsed
        self.submission_repository.update(submission)

        return ReaderInventory(
            submission_id=submission.submission_id,
            submission_intent=intent,
            document_inventory=inventory,
            warnings=warnings,
            confidence=0.84,
        )

    def _classify_intent(self, submission: SubmissionEnvelope) -> SubmissionIntent:
        text = f"{submission.subject}\n{submission.email_body_text}".lower()
        if "rfp" in text or "quote" in text or "submission" in text:
            return SubmissionIntent.rfp_submission
        if any(item.file_name.lower().endswith((".csv", ".xlsx", ".xls", ".pdf")) for item in submission.attachments):
            return SubmissionIntent.census_only
        return SubmissionIntent.general_email

    def _classify_attachment(self, attachment: Attachment) -> Attachment:
        file_name = attachment.file_name.lower()
        content_hint = self._read_content_hint(attachment)
        combined_hint = f"{file_name}\n{content_hint}".lower()
        tags: list[str] = []
        document_type = DocumentType.unknown
        if file_name.endswith((".csv", ".xlsx", ".xls")):
            if any(token in combined_hint for token in ["employee_id", "salary", "dependent_count", "census"]):
                document_type = DocumentType.census
            else:
                document_type = DocumentType.plan_summary
            tags.extend(["structured", "tabular"])
        elif file_name.endswith(".pdf"):
            if any(token in combined_hint for token in ["employee_id", "salary", "dependent_count", "census"]):
                document_type = DocumentType.census
                tags.extend(["pdf", "tabular"])
            elif any(token in combined_hint for token in ["plan", "benefit", "design", "ppo", "copay", "critical illness", "accident", "hospital indemnity"]):
                document_type = DocumentType.plan_summary
                tags.extend(["pdf", "benefits"])
            elif "rate" in combined_hint:
                document_type = DocumentType.rate_exhibit
                tags.extend(["pdf", "pricing"])
            else:
                document_type = DocumentType.narrative_rfp
                tags.extend(["pdf", "narrative"])
        elif file_name.endswith((".png", ".jpg", ".jpeg", ".tif", ".tiff")):
            document_type = DocumentType.image_scan
            tags.extend(["image", "scan"])
        attachment.document_type = document_type
        attachment.tags = sorted(set(tags))
        return attachment

    def _read_content_hint(self, attachment: Attachment) -> str:
        if not attachment.storage_key:
            return ""
        try:
            content = self.object_store.get_bytes(attachment.storage_key)
        except Exception:
            return ""
        file_name = attachment.file_name.lower()
        if file_name.endswith(".csv"):
            return content.decode("utf-8-sig", errors="ignore")[:500]
        if file_name.endswith((".xlsx", ".xls")):
            workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            parts: list[str] = []
            for sheet in workbook.worksheets[:1]:
                for row in sheet.iter_rows(values_only=True, max_row=5):
                    values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                    if values:
                        parts.append(" ".join(values))
            return "\n".join(parts)[:500]
        if file_name.endswith(".pdf"):
            return "\n".join(text for _, text in extract_pdf_text(content) if text)[:500]
        return ""
