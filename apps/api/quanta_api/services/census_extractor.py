from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from statistics import mean, median

from openpyxl import load_workbook

from quanta_api.domain.enums import SourceType
from quanta_api.domain.models import Attachment, CensusDataset, CensusSummary, EvidenceReference, FieldExtractionResult
from quanta_api.domain.repositories import ObjectStore
from quanta_api.services.id_factory import IdFactory
from quanta_api.services.ocr import ocr_pdf_rows
from quanta_api.services.pdf_tables import extract_pdf_rows, extract_pdf_text

AGE_COLUMNS = {"age", "employee_age"}
DOB_COLUMNS = {"dob", "date_of_birth", "birth_date"}
SALARY_COLUMNS = {"salary", "annual_salary", "earnings", "annual_earnings", "annualincome", "annual_income"}
STATE_COLUMNS = {"state", "employee_state", "work_state", "resident_state"}
ZIP_COLUMNS = {"zip", "zip_code", "zipcode", "postal_code"}
CLASS_COLUMNS = {"class", "employee_class", "benefit_class", "benefitclassname", "benefit_class_name"}
DEPENDENT_COLUMNS = {"dependent_count", "dependents"}
EMPLOYEE_ID_COLUMNS = {"employee_id", "employeeid", "employee_code", "employeecode", "row_id"}
HEADER_SYNONYMS = {
    "employee_id": EMPLOYEE_ID_COLUMNS,
    "age": AGE_COLUMNS,
    "birth_date": DOB_COLUMNS,
    "salary": SALARY_COLUMNS,
    "state": STATE_COLUMNS,
    "zip": ZIP_COLUMNS,
    "class": CLASS_COLUMNS,
    "dependent_count": DEPENDENT_COLUMNS,
}


class CensusExtractionService:
    def __init__(self, object_store: ObjectStore, ids: IdFactory) -> None:
        self.object_store = object_store
        self.ids = ids

    def extract(self, parent_case_id: str, attachments: list[Attachment]) -> CensusDataset:
        best_dataset: CensusDataset | None = None
        for attachment in attachments:
            if not attachment.storage_key:
                continue
            lower = attachment.file_name.lower()
            if lower.endswith(".csv"):
                rows = self._read_csv(attachment)
                dataset = self._build_dataset(parent_case_id, attachment, rows, attachment.file_name)
                best_dataset = self._pick_better(best_dataset, dataset)
            if lower.endswith(".xlsx"):
                rows, sheet_name = self._read_xlsx(attachment)
                dataset = self._build_dataset(parent_case_id, attachment, rows, attachment.file_name, sheet_name)
                best_dataset = self._pick_better(best_dataset, dataset)
            if lower.endswith(".pdf"):
                pdf_dataset = self._read_pdf(parent_case_id, attachment)
                best_dataset = self._pick_better(best_dataset, pdf_dataset)

        if best_dataset is not None:
            return best_dataset

        return CensusDataset(
            census_id=self.ids.next_census_id(),
            parent_case_id=parent_case_id,
            source_files=[item.file_name for item in attachments],
            anomalies=["No CSV/XLSX/PDF census attachment detected"],
            extraction_confidence=0.1,
        )

    def _read_csv(self, attachment: Attachment) -> list[dict[str, str]]:
        content = self.object_store.get_bytes(attachment.storage_key)
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        rows: list[dict[str, str]] = []
        for row in reader:
            normalized_row = {(key or "").strip(): self._stringify(value) for key, value in row.items()}
            if any(normalized_row.values()):
                rows.append(normalized_row)
        return rows

    def _read_xlsx(self, attachment: Attachment) -> tuple[list[dict[str, str]], str]:
        content = self.object_store.get_bytes(attachment.storage_key)
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        all_rows: list[dict[str, str]] = []
        chosen_sheet = workbook.active.title
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            header_index = self._detect_header_row(rows)
            if header_index is None:
                continue
            headers = [self._normalize_header(cell) for cell in rows[header_index]]
            sheet_rows: list[dict[str, str]] = []
            for row in rows[header_index + 1:]:
                item = {}
                for index, value in enumerate(row):
                    header = headers[index] if index < len(headers) and headers[index] else f"column_{index + 1}"
                    item[header] = self._stringify(value)
                if any(item.values()):
                    sheet_rows.append(item)
            if len(sheet_rows) > len(all_rows):
                chosen_sheet = sheet.title
            all_rows.extend(sheet_rows)
        return all_rows, chosen_sheet

    def _build_dataset(
        self,
        parent_case_id: str,
        attachment: Attachment,
        rows: list[dict[str, str]],
        file_name: str,
        sheet_name: str | None = None,
    ) -> CensusDataset:
        normalized_rows = self._normalize_rows(rows)
        columns = list(normalized_rows[0].keys()) if normalized_rows else []
        employee_count = len(normalized_rows)
        dependent_count = sum(self._safe_int(row.get("dependent_count")) for row in normalized_rows)
        classes_detected = sorted({row.get("class", "") for row in normalized_rows if row.get("class")})
        states_detected = sorted({row.get("state", "") for row in normalized_rows if row.get("state")})
        ages = [self._safe_float(row.get("age")) for row in normalized_rows if self._safe_float(row.get("age")) is not None]
        salaries = [self._safe_float(row.get("salary")) for row in normalized_rows if self._safe_float(row.get("salary")) is not None]

        evidence = [
            EvidenceReference(
                source_type=SourceType.attachment,
                file_name=file_name,
                sheet_name=sheet_name,
                cell_range=f"A1:{self._column_letter(len(columns))}{min(employee_count + 1, 25)}" if columns else None,
                snippet=f"Detected {employee_count} census rows with columns: {', '.join(columns[:8])}",
                confidence=0.92 if employee_count else 0.4,
            )
        ]

        anomalies = []
        if not employee_count:
            anomalies.append("Attachment was readable but produced zero census rows")
        if "salary" not in columns:
            anomalies.append("Salary column not detected")
        if "age" not in columns and "birth_date" not in columns:
            anomalies.append("Age or birth date column not detected")

        field_results = {
            "employee_count": FieldExtractionResult(value=employee_count, confidence=0.95 if employee_count else 0.2, evidence=evidence, warnings=[]),
            "dependent_count": FieldExtractionResult(value=dependent_count, confidence=0.9 if columns else 0.2, evidence=evidence, warnings=[]),
            "classes_detected": FieldExtractionResult(value=classes_detected, confidence=0.85 if classes_detected else 0.2, evidence=evidence, warnings=[]),
            "states_detected": FieldExtractionResult(value=states_detected, confidence=0.85 if states_detected else 0.2, evidence=evidence, warnings=[]),
            "age_metric": FieldExtractionResult(value=round(mean(ages), 1) if ages else None, confidence=0.82 if ages else 0.15, evidence=evidence, warnings=[]),
            "salary_metric": FieldExtractionResult(value=round(median(salaries), 2) if salaries else None, confidence=0.82 if salaries else 0.15, evidence=evidence, warnings=[]),
        }

        return CensusDataset(
            census_id=self.ids.next_census_id(),
            parent_case_id=parent_case_id,
            source_files=[file_name],
            employee_count=employee_count,
            dependent_count=dependent_count,
            classes_detected=classes_detected,
            states_detected=states_detected,
            census_columns_detected=columns,
            census_rows=normalized_rows[:200],
            summary_statistics=CensusSummary(avg_age=round(mean(ages), 1) if ages else None, median_salary=round(median(salaries), 2) if salaries else None),
            anomalies=anomalies,
            extraction_confidence=0.92 if employee_count else 0.3,
            evidence_references=evidence,
            field_results=field_results,
        )

    def _read_pdf(self, parent_case_id: str, attachment: Attachment) -> CensusDataset:
        content = self.object_store.get_bytes(attachment.storage_key)
        table_pages = extract_pdf_rows(content)
        text_pages = extract_pdf_text(content)
        ocr_pages: list[tuple[int, list[dict[str, str]]]] = []
        ocr_warnings: list[str] = []
        if not table_pages:
            ocr_pages, ocr_warnings = ocr_pdf_rows(content)

        normalized_rows: list[dict[str, str]] = []
        evidence: list[EvidenceReference] = []
        for page_number, rows in table_pages:
            page_rows = self._normalize_rows(rows)
            if page_rows:
                normalized_rows.extend(page_rows)
                columns = list(page_rows[0].keys())
                evidence.append(
                    EvidenceReference(
                        source_type=SourceType.attachment,
                        file_name=attachment.file_name,
                        page_number=page_number,
                        cell_range=f"A1:{self._column_letter(len(columns))}{min(len(page_rows) + 1, 25)}",
                        snippet=f"PDF table rows detected on page {page_number}: {', '.join(columns[:8])}",
                        confidence=0.86,
                    )
                )
        for page_number, rows in ocr_pages:
            page_rows = self._normalize_rows(rows)
            if page_rows:
                normalized_rows.extend(page_rows)
                evidence.append(
                    EvidenceReference(
                        source_type=SourceType.attachment,
                        file_name=attachment.file_name,
                        page_number=page_number,
                        snippet=f"OCR recovered census-like rows from scanned PDF page {page_number}",
                        confidence=0.58,
                    )
                )

        text_hint = next((text for _, text in text_pages if text), "")
        dataset = self._build_dataset(parent_case_id=parent_case_id, attachment=attachment, rows=normalized_rows, file_name=attachment.file_name)
        if evidence:
            dataset.evidence_references = evidence
            for field_name, result in dataset.field_results.items():
                result.evidence = evidence
                if any(ref.confidence < 0.7 for ref in evidence):
                    result.confidence = round(min(result.confidence, 0.68), 2)
                    result.warnings.append("Field confidence reduced because OCR-derived evidence was used")
                dataset.field_results[field_name] = result
        if ocr_warnings:
            dataset.anomalies.extend(ocr_warnings)
        if text_hint and not evidence:
            dataset.evidence_references = [
                EvidenceReference(
                    source_type=SourceType.attachment,
                    file_name=attachment.file_name,
                    page_number=text_pages[0][0] if text_pages else None,
                    snippet=text_hint[:180],
                    confidence=0.45,
                )
            ]
            dataset.anomalies.append("PDF text detected but no structured census table was extracted")
            dataset.extraction_confidence = 0.25
        return dataset

    def _pick_better(self, current: CensusDataset | None, candidate: CensusDataset) -> CensusDataset:
        if current is None:
            return candidate
        if candidate.employee_count > current.employee_count:
            return candidate
        if candidate.employee_count == current.employee_count and candidate.extraction_confidence > current.extraction_confidence:
            return candidate
        return current

    def _normalize_rows(self, rows: list[dict[str, str]]) -> list[dict[str, str]]:
        normalized_rows: list[dict[str, str]] = []
        seen_keys: set[str] = set()
        for row in rows:
            normalized = self._normalize_row(row)
            if not any(normalized.values()):
                continue
            identity = "|".join(str(normalized.get(key, "")) for key in ["employee_id", "first_name", "last_name", "birth_date", "salary"])
            if identity in seen_keys:
                continue
            seen_keys.add(identity)
            normalized_rows.append(normalized)
        return normalized_rows

    def _normalize_row(self, row: dict[str, str]) -> dict[str, str]:
        normalized: dict[str, str] = {}
        for raw_key, value in row.items():
            key = self._normalize_header(raw_key)
            if key in AGE_COLUMNS:
                normalized["age"] = self._normalize_number_string(value)
            elif key in DOB_COLUMNS:
                normalized["birth_date"] = self._normalize_date_string(value)
                if normalized.get("birth_date") and "age" not in normalized:
                    age = self._age_from_birth_date(normalized["birth_date"])
                    if age is not None:
                        normalized["age"] = str(age)
            elif key in SALARY_COLUMNS:
                normalized["salary"] = self._normalize_number_string(value)
            elif key in STATE_COLUMNS:
                normalized["state"] = self._normalize_state(value)
            elif key in ZIP_COLUMNS:
                normalized["zip"] = re.sub(r"[^0-9]", "", value)[:5]
            elif key in CLASS_COLUMNS:
                normalized["class"] = value.strip()
            elif key in DEPENDENT_COLUMNS:
                normalized["dependent_count"] = self._normalize_number_string(value)
            elif key in EMPLOYEE_ID_COLUMNS:
                normalized["employee_id"] = value.strip()
            else:
                normalized[key] = value.strip()
        if "employee_id" not in normalized:
            normalized["employee_id"] = normalized.get("row_id", "")
        if "dependent_count" not in normalized:
            normalized["dependent_count"] = "0"
        return normalized

    def _detect_header_row(self, rows: list[tuple]) -> int | None:
        for index, row in enumerate(rows[:5]):
            headers = [self._normalize_header(cell) for cell in row]
            score = sum(1 for header in headers if any(header in synonyms for synonyms in HEADER_SYNONYMS.values()))
            if score >= 2:
                return index
        return 0 if rows else None

    def _normalize_header(self, value) -> str:
        text = self._stringify(value).lower().strip()
        text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
        for normalized, synonyms in HEADER_SYNONYMS.items():
            if text in synonyms:
                return normalized
        return text

    def _normalize_number_string(self, value: str) -> str:
        return re.sub(r"[^0-9.-]", "", value or "")

    def _normalize_date_string(self, value: str) -> str:
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(value.strip(), fmt).date().isoformat()
            except (ValueError, AttributeError):
                continue
        return value.strip()

    def _normalize_state(self, value: str) -> str:
        cleaned = value.strip().upper()
        return cleaned[:2]

    def _age_from_birth_date(self, value: str) -> int | None:
        try:
            birth_date = datetime.fromisoformat(value).date()
        except ValueError:
            return None
        today = datetime.utcnow().date()
        return today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

    def _stringify(self, value) -> str:
        if isinstance(value, list):
            return " ".join(item.strip() for item in value if item and str(item).strip())
        return "" if value is None else str(value).strip()

    def _safe_float(self, value: str | None) -> float | None:
        if value in (None, ""):
            return None
        try:
            return float(value)
        except ValueError:
            return None

    def _safe_int(self, value: str | None) -> int:
        if value in (None, ""):
            return 0
        try:
            return int(float(value))
        except ValueError:
            return 0

    def _column_letter(self, column_number: int) -> str:
        if column_number <= 0:
            return "A"
        result = ""
        current = column_number
        while current:
            current, remainder = divmod(current - 1, 26)
            result = chr(65 + remainder) + result
        return result
