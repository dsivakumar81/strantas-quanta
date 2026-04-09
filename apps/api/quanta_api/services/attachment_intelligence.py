from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook

from quanta_api.domain.enums import SourceType
from quanta_api.domain.models import Attachment, EvidenceReference, FieldExtractionResult, PlanDesign
from quanta_api.domain.repositories import ObjectStore
from quanta_api.services.pdf_tables import extract_pdf_text

LOB_PATTERNS = {
    "group_life": [r"\bbasic life\b", r"\bvoluntary life\b", r"\blife\b"],
    "group_std": [r"\bstd\b", r"\bshort[- ]term disability\b"],
    "group_ltd": [r"\bltd\b", r"\blong[- ]term disability\b"],
    "dental": [r"\bdental\b", r"\bppo dental\b", r"\bdppo\b"],
    "vision": [r"\bvision\b"],
    "supplemental_ci": [r"\bcritical illness\b", r"\bcritical care\b", r"\bci\b"],
    "supplemental_accident": [r"\baccident\b"],
    "supplemental_hi": [r"\bhospital indemnity\b", r"\binpatient indemnity\b", r"\bhi\b"],
}


@dataclass
class AttachmentInsight:
    detected_lobs: set[str] = field(default_factory=set)
    lob_evidence: dict[str, list[EvidenceReference]] = field(default_factory=dict)
    plan_designs: dict[str, list[PlanDesign]] = field(default_factory=dict)
    field_candidates: dict[str, list[tuple[str, EvidenceReference]]] = field(default_factory=dict)
    lob_metadata: dict[str, dict[str, list[str]]] = field(default_factory=dict)


class AttachmentIntelligenceService:
    def __init__(self, object_store: ObjectStore) -> None:
        self.object_store = object_store

    def analyze(self, attachments: list[Attachment]) -> AttachmentInsight:
        insight = AttachmentInsight()
        for attachment in attachments:
            if not attachment.storage_key:
                continue
            raw_text = self._read_text(attachment)
            text = self._normalize_extraction_text(raw_text)
            if not text:
                continue
            for lob_type, patterns in LOB_PATTERNS.items():
                matches = [pattern for pattern in patterns if re.search(pattern, text, re.IGNORECASE)]
                if matches:
                    insight.detected_lobs.add(lob_type)
                    if lob_type == "dental":
                        snippet = self._clean_snippet(
                            self._extract_section(
                                text,
                                ["dental plan", "requested dental plan"],
                                ["vision plan", "census", "critical illness", "accident", "hospital indemnity"],
                            )
                        )
                    elif lob_type == "vision":
                        snippet = self._clean_snippet(
                            self._extract_section(
                                text,
                                ["vision plan", "requested vision plan"],
                                ["census", "critical illness", "accident", "hospital indemnity"],
                            )
                        )
                    else:
                        snippet = self._matching_snippet(text, matches[0])
                    evidence = EvidenceReference(
                        source_type=SourceType.attachment,
                        file_name=attachment.file_name,
                        snippet=snippet,
                        confidence=0.78,
                    )
                    insight.lob_evidence.setdefault(lob_type, []).append(evidence)
                    plans = self._plan_designs_for(lob_type, text, attachment.file_name, evidence)
                    if plans:
                        insight.plan_designs.setdefault(lob_type, []).extend(plans)
                    self._collect_lob_metadata(insight, lob_type, text)
            self._collect_field_candidates(insight, attachment.file_name, text)
        return insight

    def _read_text(self, attachment: Attachment) -> str:
        content = self.object_store.get_bytes(attachment.storage_key)
        file_name = attachment.file_name.lower()
        if file_name.endswith(".csv"):
            return content.decode("utf-8-sig", errors="ignore")
        if file_name.endswith(".xlsx"):
            workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            chunks: list[str] = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                    if values:
                        chunks.append(" ".join(values))
            return "\n".join(chunks)
        if file_name.endswith(".pdf"):
            return "\n".join(text for _, text in extract_pdf_text(content) if text)
        return ""

    def _plan_designs_for(self, lob_type: str, text: str, file_name: str, evidence: EvidenceReference) -> list[PlanDesign]:
        if lob_type == "group_life":
            basis_match = re.search(r"(\d+)\s*[x×]\s*(?:annual )?salary", text, re.IGNORECASE)
            max_match = re.search(r"max(?:imum)?\s+(?:benefit|life benefit)\s*[:\-]?\s*\$?([\d,]+)", text, re.IGNORECASE)
            gi_match = re.search(r"guarantee(?:d)? issue\s*[:\-]?\s*\$?([\d,]+)", text, re.IGNORECASE)
            if basis_match or max_match:
                return [self._plan(
                    plan_type="basic_life",
                    notes=[f"Attachment-derived life design from {file_name}"],
                    evidence=evidence,
                    benefit_basis=f"{basis_match.group(1)}x_salary" if basis_match else None,
                    max_benefit=float(max_match.group(1).replace(",", "")) if max_match else None,
                    guarantee_issue=float(gi_match.group(1).replace(",", "")) if gi_match else None,
                )]
        if lob_type in {"group_ltd", "group_std"}:
            percent_match = re.search(
                r"(?:benefit|replaces?|pays?)\s*(\d{1,3})\s*%|(?<!\d)(\d{1,3})\s*%\s*(?:of\s+covered earnings|of earnings|of salary|benefit|covered earnings)",
                text,
                re.IGNORECASE,
            )
            elimination_match = re.search(r"(?:elimination period|waiting period|elim period)\s*(?:of)?\s*(\d+)\s*days", text, re.IGNORECASE)
            max_match = re.search(r"max(?:imum)?\s+(monthly|weekly)\s+benefit\s*[:\-]?\s*\$?([\d,]+)", text, re.IGNORECASE)
            duration_match = re.search(r"benefit duration\s*[:\-]?\s*([a-z0-9 ,/\-]+)", text, re.IGNORECASE)
            contribution_match = re.search(r"employer[_ ]paid|employee[_ ]paid|voluntary|contributory", text, re.IGNORECASE)
            hours_match = re.search(r"eligible employees working\s*(\d+)\+?\s*hours", text, re.IGNORECASE)
            percent_groups = [group for group in percent_match.groups() if group] if percent_match else []
            percent_value = float(percent_groups[0]) if percent_groups else None
            if percent_value or elimination_match or max_match:
                return [self._plan(
                    plan_type="ltd" if lob_type == "group_ltd" else "std",
                    notes=[f"Attachment-derived disability design from {file_name}"],
                    evidence=evidence,
                    benefit_percent=percent_value,
                    elimination_period_days=int(elimination_match.group(1)) if elimination_match else None,
                    max_monthly_benefit=float(max_match.group(2).replace(",", "")) if max_match and max_match.group(1).lower() == "monthly" else None,
                    max_weekly_benefit=float(max_match.group(2).replace(",", "")) if max_match and max_match.group(1).lower() == "weekly" else None,
                    contribution_details=contribution_match.group(0).replace("_", " ") if contribution_match else None,
                    attributes={key: value for key, value in {
                        "benefit_duration": duration_match.group(1).strip() if duration_match else None,
                        "min_weekly_eligible_hours": int(hours_match.group(1)) if hours_match else None,
                    }.items() if value is not None},
                )]
        if lob_type == "dental":
            designs: list[PlanDesign] = []
            section_text = self._extract_section(text, ["dental plan", "requested dental plan"], ["vision plan", "census", "critical illness", "accident", "hospital indemnity"])
            if file_name.lower().endswith(".pdf"):
                segments = [section_text or text]
            else:
                segments = self._extract_segments(section_text or text, ["dental"])
            for segment in segments:
                coverage_match = re.search(r"(100\s*/\s*80\s*/\s*50|90\s*/\s*80\s*/\s*50|80\s*/\s*80\s*/\s*50)", segment, re.IGNORECASE)
                contribution_match = re.search(r"employer[_ ]paid|employee[_ ]paid|contributory|voluntary", segment, re.IGNORECASE)
                ortho_match = re.search(r"ortho(?:dontia)?\s+(\d+)%", segment, re.IGNORECASE)
                ortho_age_match = re.search(r"ortho(?:dontia)?.{0,20}?(?:to age|age limit)\s*(\d{1,2})", segment, re.IGNORECASE)
                deductible_match = re.search(r"deductible\s*[:\-]?\s*\$?(\d+)", segment, re.IGNORECASE)
                annual_max_match = re.search(r"annual max(?:imum)?\s*[:\-]?\s*\$?([\d,]+)", segment, re.IGNORECASE)
                preventive_match = re.search(r"preventive\s*(\d+)%", segment, re.IGNORECASE)
                basic_match = re.search(r"basic\s*(\d+)%", segment, re.IGNORECASE)
                major_match = re.search(r"major\s*(\d+)%", segment, re.IGNORECASE)
                office_visit_match = re.search(r"(?:office visit|diagnostic visit)\s+copay\s*\$?(\d+)", segment, re.IGNORECASE)
                waiting_match = re.search(r"service waiting periods?\s*[:\-]?\s*([a-z0-9 ,/\-]+)", segment, re.IGNORECASE)
                if re.search(r"\bppo\b|\bdhmo\b|\bepo\b", segment, re.IGNORECASE) or coverage_match or contribution_match:
                    notes = [f"Attachment-derived dental design from {file_name}"]
                    designs.append(self._plan(
                        plan_type="dhmo" if re.search(r"\bdhmo\b", segment, re.IGNORECASE) else "ppo" if re.search(r"\bppo\b|\bepo\b", segment, re.IGNORECASE) else "dental",
                        notes=notes,
                        evidence=evidence,
                        contribution_details=contribution_match.group(0).replace("_", " ") if contribution_match else None,
                        attributes={key: value for key, value in {
                            "coverage_tiers": coverage_match.group(1).replace(" ", "") if coverage_match else None,
                            "preventive_percent": int(preventive_match.group(1)) if preventive_match else None,
                            "basic_percent": int(basic_match.group(1)) if basic_match else None,
                            "major_percent": int(major_match.group(1)) if major_match else None,
                            "orthodontia_percent": int(ortho_match.group(1)) if ortho_match else None,
                            "orthodontia_age_limit": int(ortho_age_match.group(1)) if ortho_age_match else None,
                            "deductible": int(deductible_match.group(1)) if deductible_match else None,
                            "annual_maximum": float(annual_max_match.group(1).replace(",", "")) if annual_max_match else None,
                            "office_visit_copay": int(office_visit_match.group(1)) if office_visit_match else None,
                            "service_waiting_periods": waiting_match.group(1).strip() if waiting_match else None,
                        }.items() if value is not None},
                    ))
            return designs
        if lob_type == "vision":
            section_text = self._extract_section(text, ["vision plan", "requested vision plan"], ["census", "critical illness", "accident", "hospital indemnity"])
            exam_match = re.search(r"exam\s+copay\s+\$?(\d+)", section_text or text, re.IGNORECASE)
            material_match = re.search(r"materials?\s+copay\s+\$?(\d+)", section_text or text, re.IGNORECASE)
            frame_match = re.search(r"frames?\s+(?:allowance|benefit)\s*[:\-]?\s*\$?([\d,]+)", section_text or text, re.IGNORECASE)
            contact_match = re.search(r"contacts?\s+(?:allowance|benefit)\s*[:\-]?\s*\$?([\d,]+)", section_text or text, re.IGNORECASE)
            frequency_match = re.search(r"every\s+(\d+)\s+months", section_text or text, re.IGNORECASE)
            lens_match = re.search(r"(?:lenses?|lens materials?)\s+copay\s+\$?(\d+)", section_text or text, re.IGNORECASE)
            laser_match = re.search(r"laser vision correction\s*[:\-]?\s*\$?([\d,]+)", section_text or text, re.IGNORECASE)
            if re.search(r"\bvision\b", section_text or text, re.IGNORECASE):
                return [self._plan(
                    plan_type="vision",
                    notes=[f"Attachment-derived vision design from {file_name}"],
                    evidence=evidence,
                    attributes={key: value for key, value in {
                        "exam_copay": int(exam_match.group(1)) if exam_match else None,
                        "materials_copay": int(material_match.group(1)) if material_match else None,
                        "frame_allowance": float(frame_match.group(1).replace(",", "")) if frame_match else None,
                        "contact_allowance": float(contact_match.group(1).replace(",", "")) if contact_match else None,
                        "frequency_months": int(frequency_match.group(1)) if frequency_match else None,
                        "lens_copay": int(lens_match.group(1)) if lens_match else None,
                        "laser_correction_allowance": float(laser_match.group(1).replace(",", "")) if laser_match else None,
                    }.items() if value is not None},
                )]
        if lob_type == "supplemental_ci":
            benefit_match = re.search(r"(?:critical illness|ci)(?:[^$\d]{0,40})\$?([\d,]+)", text, re.IGNORECASE)
            gi_match = re.search(r"guarantee(?:d)? issue\s*[:\-]?\s*\$?([\d,]+)", text, re.IGNORECASE)
            wellness_match = re.search(r"wellness(?: benefit)?\s*[:\-]?\s*\$?([\d,]+)", text, re.IGNORECASE)
            if benefit_match:
                return [self._plan(
                    plan_type="critical_illness",
                    notes=[f"Attachment-derived CI design from {file_name}"],
                    evidence=evidence,
                    max_benefit=float(benefit_match.group(1).replace(",", "")),
                    guarantee_issue=float(gi_match.group(1).replace(",", "")) if gi_match else None,
                    attributes={"wellness_benefit": float(wellness_match.group(1).replace(",", ""))} if wellness_match else {},
                )]
        if lob_type == "supplemental_accident":
            off_job_match = re.search(r"off[- ]job(?: benefit)?\s*[:\-]?\s*\$?([\d,]+)", text, re.IGNORECASE)
            er_match = re.search(r"er visit(?: benefit)?\s*[:\-]?\s*\$?([\d,]+)", text, re.IGNORECASE)
            if re.search(r"accident", text, re.IGNORECASE):
                return [self._plan(
                    plan_type="accident",
                    notes=[f"Attachment-derived accident design from {file_name}"],
                    evidence=evidence,
                    attributes={key: value for key, value in {
                        "off_job_benefit": float(off_job_match.group(1).replace(",", "")) if off_job_match else None,
                        "er_visit_benefit": float(er_match.group(1).replace(",", "")) if er_match else None,
                    }.items() if value is not None},
                )]
        if lob_type == "supplemental_hi":
            benefit_match = re.search(r"hospital indemnity(?:[^$\d]{0,40})\$?([\d,]+)", text, re.IGNORECASE)
            admission_match = re.search(r"admission(?: benefit)?\s*[:\-]?\s*\$?([\d,]+)", text, re.IGNORECASE)
            confinement_match = re.search(r"daily confinement(?: benefit)?\s*[:\-]?\s*\$?([\d,]+)", text, re.IGNORECASE)
            if benefit_match:
                return [self._plan(
                    plan_type="hospital_indemnity",
                    notes=[f"Attachment-derived hospital indemnity design from {file_name}"],
                    evidence=evidence,
                    max_benefit=float(benefit_match.group(1).replace(",", "")),
                    attributes={key: value for key, value in {
                        "admission_benefit": float(admission_match.group(1).replace(",", "")) if admission_match else None,
                        "daily_confinement_benefit": float(confinement_match.group(1).replace(",", "")) if confinement_match else None,
                    }.items() if value is not None},
                )]
        return []

    def _plan(self, evidence: EvidenceReference, notes: list[str], **kwargs) -> PlanDesign:
        plan = PlanDesign(notes=notes, **kwargs)
        field_results = {}
        for field_name in [
            "plan_type", "benefit_basis", "benefit_percent", "elimination_period_days", "max_benefit", "max_monthly_benefit", "max_weekly_benefit", "guarantee_issue", "contribution_details"
        ]:
            value = getattr(plan, field_name)
            if value is not None:
                field_results[field_name] = FieldExtractionResult(
                    value=value,
                    confidence=self._field_confidence(field_name, evidence),
                    evidence=[evidence],
                    warnings=[],
                )
        for key, value in plan.attributes.items():
            field_results[key] = FieldExtractionResult(
                value=value,
                confidence=self._field_confidence(key, evidence),
                evidence=[evidence],
                warnings=[],
            )
        plan.field_results = field_results
        return plan

    def _normalize_extraction_text(self, text: str) -> str:
        normalized = text.lower()
        normalized = normalized.replace("|", " ").replace("_", " ")
        normalized = normalized.replace("1ife", "life").replace("vlsion", "vision")
        normalized = normalized.replace("critica1", "critical").replace("hospita1", "hospital")
        normalized = normalized.replace("acc1dent", "accident").replace("denta1", "dental")
        normalized = normalized.replace("o%", "0%")
        normalized = re.sub(r"[\u2010-\u2015]", "-", normalized)
        normalized = re.sub(r"(?<=\d)\s+(?=%)", "", normalized)
        normalized = re.sub(r"(?<=\$)\s+", "", normalized)
        normalized = normalized.replace("employee paid", "employee_paid").replace("employer paid", "employer_paid")
        normalized = re.sub(r"\s+", " ", normalized)
        return normalized

    def _field_confidence(self, field_name: str, evidence: EvidenceReference) -> float:
        base = evidence.confidence
        if field_name in {"benefit_percent", "max_benefit", "max_monthly_benefit", "max_weekly_benefit", "guarantee_issue"}:
            base += 0.06
        elif field_name in {"elimination_period_days", "exam_copay", "materials_copay", "deductible", "annual_maximum"}:
            base += 0.04
        elif field_name in {"plan_type", "benefit_basis", "contribution_details"}:
            base += 0.02
        file_name = (evidence.file_name or "").lower()
        if file_name.endswith(".xlsx"):
            base += 0.04
        elif file_name.endswith(".pdf"):
            base += 0.0
        elif file_name.endswith(".csv"):
            base -= 0.02
        return round(min(max(base, 0.0), 0.99), 2)

    def _matching_snippet(self, text: str, pattern: str) -> str:
        match = re.search(pattern, text, re.IGNORECASE)
        if not match:
            return self._clean_snippet(text[:180])
        sentence_start = max(text.rfind(".", 0, match.start()), text.rfind("\n", 0, match.start()))
        sentence_end_candidates = [candidate for candidate in (text.find(".", match.end()), text.find("\n", match.end())) if candidate != -1]
        sentence_end = min(sentence_end_candidates) if sentence_end_candidates else len(text)
        start = max(0, sentence_start + 1 if sentence_start != -1 else match.start() - 50)
        end = min(len(text), sentence_end + 1 if sentence_end != len(text) else match.end() + 120)
        return self._clean_snippet(text[start:end])

    def _clean_snippet(self, text: str, max_length: int | None = 220) -> str:
        cleaned = text.replace("\n", " ").replace("_", " ")
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
        return cleaned[:max_length] if max_length is not None else cleaned

    def _extract_segments(self, text: str, anchor_terms: list[str]) -> list[str]:
        joined = "|".join(anchor_terms)
        segments = re.findall(rf"(?:{joined})[^.]+(?:\.)?", text, re.IGNORECASE)
        return segments or [text]

    def _extract_section(self, text: str, start_terms: list[str], end_terms: list[str]) -> str:
        start_pattern = "|".join(re.escape(term) for term in start_terms)
        end_pattern = "|".join(re.escape(term) for term in end_terms)
        match = re.search(rf"({start_pattern})(.*?)(?={end_pattern}|$)", text, re.IGNORECASE | re.DOTALL)
        if not match:
            return text
        return self._clean_snippet(match.group(0), max_length=None)

    def _collect_lob_metadata(self, insight: AttachmentInsight, lob_type: str, text: str) -> None:
        metadata = insight.lob_metadata.setdefault(
            lob_type,
            {
                "class_structure": [],
                "eligibility_rules": [],
                "participation_minimums": [],
                "contribution_splits": [],
                "waiting_periods": [],
            },
        )
        class_matches = re.findall(
            r"class\s*\d+\s*[:\-]?\s*([a-z0-9 ,/&-]+?)(?=\s+(?:class\s*\d+|eligibility|employer pays|employee pays|participation|waiting period|service waiting periods|$))",
            text,
            re.IGNORECASE,
        )
        class_keyword_matches = re.findall(
            r"\b(executive|executives|salaried|hourly|union|non-union)\b",
            text,
            re.IGNORECASE,
        )
        eligibility_matches = re.findall(
            r"("
            r"all active [a-z -]+ employees"
            r"|all full[- ]time employees(?: working \d+\+?\s*hours(?: per week)?)?"
            r"|all benefit eligible employees"
            r"|eligible employees working \d+\+?\s*hours(?: per week)?"
            r"|employees working \d+\+?\s*hours(?: per week)?"
            r"|date of hire"
            r"|day one coverage"
            r"|day one"
            r"|first of the month following \d+\s*days"
            r"|first day of the month following \d+\s*days"
            r"|first of month following \d+\s*days"
            r")",
            text,
            re.IGNORECASE,
        )
        participation_matches = re.findall(
            r"(participation(?: minimum)?\s*[:\-]?\s*[a-z0-9 %+/()-]+|\d+%\s+of eligible employees(?: electing coverage)?)",
            text,
            re.IGNORECASE,
        )
        contribution_matches = re.findall(
            r"("
            r"employer[_ ]paid"
            r"|employee[_ ]paid"
            r"|employer pays \d+%[^.]*"
            r"|employee pays \d+%[^.]*"
            r"|employer pays employee only[^.]*"
            r"|dependents voluntary"
            r"|employee contribution required[^.]*"
            r"|contributory"
            r"|voluntary"
            r")",
            text,
            re.IGNORECASE,
        )
        waiting_matches = re.findall(
            r"("
            r"waiting period\s*[:\-]?\s*[a-z0-9 +/-]+"
            r"|service waiting periods?\s*[:\-]?\s*[a-z0-9 ,/\-]+"
            r"|day one coverage"
            r"|day one"
            r"|elimination period\s*[:\-]?\s*\d+\s*days"
            r"|first of the month following \d+\s*days"
            r"|first day of the month following \d+\s*days"
            r")",
            text,
            re.IGNORECASE,
        )

        metadata["class_structure"].extend(item.strip() for item in class_matches)
        metadata["class_structure"].extend(item.strip() for item in class_keyword_matches)
        metadata["eligibility_rules"].extend(item.strip() for item in eligibility_matches)
        metadata["participation_minimums"].extend(item.strip() for item in participation_matches)
        metadata["contribution_splits"].extend(item.replace("_", " ").strip() for item in contribution_matches)
        metadata["waiting_periods"].extend(item.strip() for item in waiting_matches)
        for key, values in metadata.items():
            metadata[key] = sorted(set(values))

    def _collect_field_candidates(self, insight: AttachmentInsight, file_name: str, text: str) -> None:
        patterns = {
            "employer_name": r"employer\s*[:\-]\s*([A-Za-z0-9&.,' -]+?)(?=\s+(?:broker|effective date|situs|market segment|renewal|new business)\s*[:\-]|\s*$)",
            "employer_contact": r"employer contact\s*[:\-]\s*([A-Za-z0-9&.,' -]+?)(?=\s+(?:employer email|broker|effective date|situs)\s*[:\-]|\s*$)",
            "employer_contact_name": r"employer contact\s*[:\-]\s*([A-Za-z0-9&.,' -]+?)(?=\s+(?:employer email|broker|effective date|situs)\s*[:\-]|\s*$)",
            "employer_email": r"employer email\s*[:\-]\s*([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})",
            "employer_contact_email": r"employer email\s*[:\-]\s*([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})",
            "broker_name": r"broker\s*[:\-]\s*([A-Za-z0-9&.,' -]+?)(?=\s+(?:effective date|situs|market segment)\s*[:\-]|\s*$)",
            "broker_agency_name": r"broker\s*[:\-]\s*([A-Za-z0-9&.,' -]+?)(?=\s+(?:effective date|situs|market segment|broker contact|broker email)\s*[:\-]|\s*$)",
            "broker_contact": r"broker contact\s*[:\-]\s*([A-Za-z0-9&.,' -]+?)(?=\s+(?:broker email|effective date|situs)\s*[:\-]|\s*$)",
            "broker_contact_name": r"broker contact\s*[:\-]\s*([A-Za-z0-9&.,' -]+?)(?=\s+(?:broker email|effective date|situs)\s*[:\-]|\s*$)",
            "broker_email": r"broker email\s*[:\-]\s*([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})",
            "broker_contact_email": r"broker email\s*[:\-]\s*([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})",
            "effective_date": r"(?:effective date|renewal date)\s*[:\-]?\s*(20\d{2}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/20\d{2})",
            "response_due_date": r"(?:due date|response due|proposal due)\s*[:\-]?\s*(20\d{2}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/20\d{2})",
            "situs_state": r"situs(?: state)?\s*[:\-]\s*([A-Z]{2})",
            "worksite_address": r"worksite address\s*[:\-]\s*([A-Za-z0-9#.,' -]+?)(?=\s+(?:city|state|zip)\s*[:\-]|\s*$)",
            "city": r"city\s*[:\-]\s*([A-Za-z .'-]+?)(?=\s+(?:state|zip)\s*[:\-]|\s*$)",
            "postal_code": r"(?:zip|postal code)\s*[:\-]\s*(\d{5}(?:-\d{4})?)",
            "market_segment": r"market segment\s*[:\-]\s*([A-Za-z ]+?)(?=\s+(?:dental|vision|critical illness|accident|hospital indemnity)\b|\s*$)",
            "incumbent_carrier": r"incumbent(?: carrier)?\s*[:\-]\s*([A-Za-z0-9&.,' -]+?)(?=\s+(?:due date|proposal due|dental|vision|critical illness)\b|\s*$)",
            "quote_type": r"\b(renewal|new business|requote|amendment)\b",
        }
        for field_name, pattern in patterns.items():
            for match in re.finditer(pattern, text, re.IGNORECASE):
                value = next(group for group in match.groups() if group).strip()
                evidence = EvidenceReference(
                    source_type=SourceType.attachment,
                    file_name=file_name,
                    snippet=self._matching_snippet(text, pattern),
                    confidence=0.76,
                )
                insight.field_candidates.setdefault(field_name, []).append((value, evidence))
